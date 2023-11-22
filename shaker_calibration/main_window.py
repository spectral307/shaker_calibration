from PyQt6.QtWidgets import QMainWindow, QFileDialog, QApplication
from PyQt6.QtCore import QSettings
from openpyxl import load_workbook
from os.path import dirname, basename, join, splitext
from .ui.ui_main_window import Ui_MainWindow
from .settings_dialog import SettingsDialog


class MainWindow(QMainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.__ui = Ui_MainWindow()
        self.__ui.setupUi(self)

        # self.setFixedSize(self.size())

        self.__ui.open_afc_report_action.triggered.connect(
            self.process_afc_report)
        self.__ui.settings_action.triggered.connect(self.show_settings_dialog)
        self.__ui.exit_action.triggered.connect(self.exit)

    def show_settings_dialog(self):
        SettingsDialog().exec()

    def process_afc_report(self):
        settings = QSettings()

        afc_report_path = self.__get_afc_report_path(settings)
        if not afc_report_path:
            return

        afc_report_dir = dirname(afc_report_path)
        afc_report_filename = basename(afc_report_path)

        f, s = self.__read_afc_report(afc_report_path)

        extra_f = [int(val)
                   for val in settings.value("extra_f", type=list)]
        if len(extra_f) > 0:
            f, s = self.__interpolate(f, s, extra_f)

        if settings.value("transform_into_velocity_sensitivity", type=bool):
            s_velo = self.__transform_into_velocity_sensitivity(f, s)

            afcref_path = self.__get_afcref_path(settings,
                                             afc_report_dir,
                                             afc_report_filename)
            if not afcref_path:
                return

            self.__write_afcref_file(afcref_path, f, s_velo)

            self.__ui.statusbar.showMessage(f"{afcref_path} сохранен")

        if settings.value("transform_into_displacement_sensitivity", type=bool):
            s_disp = self.__transform_into_displacement_sensitivity(f, s)

            afcref_path = self.__get_afcref_path(settings,
                                             afc_report_dir,
                                             afc_report_filename)
            if not afcref_path:
                return

            self.__write_afcref_file(afcref_path, f, s_disp)

            self.__ui.statusbar.showMessage(f"{afcref_path} сохранен")

    def exit(self):
        QApplication.quit()

    def __write_afcref_file(self, path, f, s):
        with open(path, "w") as file:
            file.writelines([
                "<?xml version=\"1.0\" encoding=\"windows-1251\"?>\n",
                "<afc_ref>\n",
                "<table>\n",
                "<column width=\"100\"/>\n",
                "<column width=\"100\"/>\n"
            ])

            for i, v in enumerate(f):
                file.write(
                    f"<row frequency=\"{f[i]}\" sensitivity=\"{'{:.12f}'.format(s[i])}\"/>\n")

            file.writelines([
                "</table>\n",
                "</afc_ref>\n"
            ])

    def __get_afc_report_path(self, settings: QSettings):
        default_afc_report_dir = settings.value("default_afc_report_dir")

        res = QFileDialog.getOpenFileName(
            self, "Открыть протокол АЧХ",
            default_afc_report_dir, "Файлы xlsm (*.xlsm)")
        afc_report_path = res[0]

        if afc_report_path:
            afc_report_dir = dirname(afc_report_path)
            if afc_report_dir != default_afc_report_dir:
                settings.setValue("default_afc_report_dir", afc_report_dir)

        return afc_report_path

    def __get_afcref_path(self, settings: QSettings, afc_report_dir, afc_report_filename):
        if settings.value("use_same_default_afc_report_and_afcref_dir", type=bool):
            default_afcref_dir = afc_report_dir
        else:
            default_afcref_dir = settings.value("default_afcref_dir")

        default_afcref_path = join(default_afcref_dir, splitext(
            afc_report_filename)[0] + ".afcref")

        res = QFileDialog.getSaveFileName(
            self, "Сохранить afcref-файл",
            default_afcref_path, "Файлы afcref (*.afcref)")
        afcref_path = res[0]

        if afcref_path:
            afcref_dir = dirname(afcref_path)

            if afcref_dir == afc_report_dir:
                settings.setValue(
                    "use_same_default_afc_report_and_afcref_dir", True)
            else:
                settings.setValue(
                    "use_same_default_afc_report_and_afcref_dir", False)
                if afcref_dir != settings.value("default_afcref_dir"):
                    settings.setValue("default_afcref_dir", afcref_dir)

        return afcref_path

    def __read_afc_report(self, path):
        wb = load_workbook(path)
        ws = wb["data"]

        cur_row = 2
        cur_col = 5

        f = []
        while value := ws.cell(cur_row, cur_col).value:
            f.append(value)
            cur_row += 1

        cur_row = 2
        cur_col += 1

        s = []
        while value := ws.cell(cur_row, cur_col).value:
            s.append(value)
            cur_row += 1

        if len(f) != len(s):
            raise BaseException(
                "Длина вектора частоты не равна длине вектора чувствительности.")

        return f, s

    def __transform_into_velocity_sensitivity(self, f, s):
        from math import pi
        s_velocity = []
        for i, v in enumerate(f):
            s_velocity_value = s[i] * 2 * pi * v / 1e3
            s_velocity.append(s_velocity_value)
        return s_velocity

    def __transform_into_displacement_sensitivity(self, f, s):
        from math import pi
        s_displacement = []
        for i, v in enumerate(f):
            s_displacement_value = s[i] * 4 * pi**2 * v**2 / 1e6
            s_displacement.append(s_displacement_value)
        return s_displacement

    def __interpolate(self, f: list, s: list, extra_f: list):
        for v in extra_f:
            if v in f:
                raise ValueError(
                    f"Значение {v} уже содержится в векторе частот.")

        from scipy.interpolate import PchipInterpolator
        from bisect import bisect

        f = f.copy()
        s = s.copy()

        interpolator = PchipInterpolator(f, s)
        extra_s = interpolator(extra_f)

        for i, v in enumerate(extra_f):
            insert_pos = bisect(f, v)
            f.insert(insert_pos, v)
            s.insert(insert_pos, extra_s[i])

        return f, s
